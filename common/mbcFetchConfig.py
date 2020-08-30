from retry import retry
from openpyxl import load_workbook
from common.tenants import *
from common.auth import *
import requests
def ErrorCategory():

    igb = load_workbook(filename=r"error_Category.xlsx")

    ignore_sheet = igb["Conditions"]
    # maximum number of rows in the Error Category sheet
    ignore_row_max = ignore_sheet.max_row
    return ignore_sheet,ignore_row_max
#ignore_sheet,ignore_row_max=ErrorCategory()


@retry(tries=6,backoff=2)
def logDownload(id,value,eLog):
    global ignore_sheet,ignore_row_max

    path = ".hana.ondemand.com/api/v1/MessageProcessingLogErrorInformations('"
    e = requests.get(Tenants[id][0] + path + value[0] + "')/$value", headers=hder)
    eLog.update({value[1]: [e.text,""]})  # Appended Error Text
    print(value[1], "-->", e.text)